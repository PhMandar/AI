import imaplib
import email
import os
import time
import requests
import smtplib
import logging
from email.message import EmailMessage
from email.utils import parsedate_to_datetime
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv
from datetime import datetime, timezone
from colorama import init, Fore, Style

# Initialize colored output
init(autoreset=True)

# Load environment variables
load_dotenv()
GMAIL_USER = os.getenv("GMAIL_USER")
GMAIL_PASS = os.getenv("GMAIL_APP_PASSWORD")
TARGET_SENDER = os.getenv("TARGET_SENDER")
FILTER_START_DATE_STR = os.getenv("FILTER_START_DATE")

# Setup logging
logging.basicConfig(filename='app.log', level=logging.INFO, format='%(asctime)s %(levelname)s:%(message)s')

# Parse start date
try:
    FILTER_START_DATE = datetime.fromisoformat(FILTER_START_DATE_STR.replace("Z", "+00:00"))
except Exception as e:
    print(Fore.RED + "❌ Invalid FILTER_START_DATE in .env.")
    logging.error("Invalid FILTER_START_DATE format.")
    exit(1)

def read_last_timestamp():
    if os.path.exists("last_time.txt"):
        with open("last_time.txt", "r") as f:
            try:
                return datetime.fromisoformat(f.read().strip())
            except:
                return None
    return None

def save_last_timestamp(dt):
    with open("last_time.txt", "w") as f:
        f.write(dt.isoformat())

def fetch_email_after_date(sender_email, after_utc):
    imap = imaplib.IMAP4_SSL("imap.gmail.com")
    imap.login(GMAIL_USER, GMAIL_PASS)
    imap.select("inbox")
    _, messages = imap.search(None, f'FROM "{sender_email}"')
    email_ids = messages[0].split()

    for eid in reversed(email_ids):
        _, msg_data = imap.fetch(eid, "(RFC822)")
        msg = email.message_from_bytes(msg_data[0][1])
        try:
            email_date = parsedate_to_datetime(msg["Date"]).astimezone(timezone.utc)
        except Exception:
            continue
        if email_date <= after_utc:
            continue

        body = ""
        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_type() == "text/plain":
                    body += part.get_payload(decode=True).decode(errors="ignore")
        else:
            body = msg.get_payload(decode=True).decode(errors="ignore")

        imap.logout()
        return body, email_date

    imap.logout()
    return None, after_utc

def parse_with_llm(email_body):
    prompt = f"""
Extract user name, consumed %, remaining %, and recommend if admin should notify the user if consumption > 85%. Format:

User | Consumed Data | Remaining Data | Notify
Mandar | 87% | 13% | Yes

Email:
---
{email_body}
---
"""
    response = requests.post(
        "http://localhost:11434/api/generate",
        json={"model": "gemma3", "prompt": prompt, "stream": False}
    )

    if response.status_code != 200:
        logging.error("LLM error: " + response.text)
        return ""

    return response.json().get("response", "")

def append_to_excel(llm_output, email_time, filename="output.xlsx"):
    rows = []
    for line in llm_output.strip().splitlines():
        if "|" in line and "User" not in line:
            cells = [cell.strip() for cell in line.split("|")]
            if len(cells) >= 3:
                rows.append([email_time.strftime("%Y-%m-%d %H:%M:%S")] + cells)

    if not rows:
        return

    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "User", "Consumed Data", "Remaining Data", "Notify"])

    for row in rows:
        ws.append(row)

    wb.save(filename)
    logging.info("Appended to Excel.")

def send_alert_email(to_email, subject, body):
    msg = EmailMessage()
    msg["From"] = GMAIL_USER
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.set_content(body)

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(GMAIL_USER, GMAIL_PASS)
            smtp.send_message(msg)
        print(Fore.GREEN + f"✅ Alert email sent to {to_email}")
        logging.info(f"Alert email sent to {to_email}")
    except Exception as e:
        print(Fore.RED + f"❌ Failed to send alert email: {e}")
        logging.error("SMTP send error: " + str(e))

def handle_notification(llm_output, sender_email):
    for line in llm_output.strip().splitlines():
        if "|" in line and "User" not in line:
            parts = [part.strip() for part in line.split("|")]
            if len(parts) >= 4 and parts[3].lower() == "yes":
                print(Fore.YELLOW + f"❗ User {parts[0]} has consumed {parts[1]}. Notify: Yes")
                choice = input(Fore.CYAN + "📧 Do you want to notify the user? (yes/no): ").strip().lower()
                if choice == "yes":
                    subject = "Consumption Report from Admin"
                    body = f"Hi {parts[0]},\n\nYour data usage is {parts[1]}.\nPlease take necessary action.\n\nThanks,\nAdmin"
                    send_alert_email(sender_email, subject, body)
                else:
                    print(Fore.GREEN + f"✅ Email not sent. I will keep monitoring your inbox for new emails.")

if __name__ == "__main__":
    print(Fore.CYAN + "📧 Gmail Monitor with LLM and Excel Logging (Running until Ctrl+C)")

    last_ts = read_last_timestamp() or FILTER_START_DATE
    print(Fore.BLUE + f"🕒 Starting from: {last_ts.isoformat()}")

    try:
        while True:
            email_body, email_time = fetch_email_after_date(TARGET_SENDER, last_ts)
            if email_body:
                print(Fore.GREEN + "📨 New email found. Processing...")
                llm_output = parse_with_llm(email_body)
                print(Fore.WHITE + llm_output)

                append_to_excel(llm_output, email_time)
                handle_notification(llm_output, TARGET_SENDER)

                last_ts = email_time
                save_last_timestamp(last_ts)

            time.sleep(30)

    except KeyboardInterrupt:
        print(Fore.RED + "\n🛑 Stopped by user.")
        logging.info("Stopped by user.")