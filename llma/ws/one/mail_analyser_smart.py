import imaplib
import email
import os
import time
import requests
import smtplib
import logging
from email.message import EmailMessage
from email.utils import parsedate_to_datetime
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv
from datetime import datetime, timezone
from colorama import init, Fore, Style

# Initialize colored output
init(autoreset=True)

# Load environment variables
load_dotenv()
GMAIL_USER = os.getenv("GMAIL_USER")
GMAIL_PASS = os.getenv("GMAIL_APP_PASSWORD")
TARGET_SENDER = os.getenv("TARGET_SENDER")
FILTER_START_DATE_STR = os.getenv("FILTER_START_DATE")  # ISO format

# Setup logging
logging.basicConfig(filename='app.log', level=logging.INFO, format='%(asctime)s %(levelname)s:%(message)s')

# Parse filter start date from env
try:
    FILTER_START_DATE = datetime.fromisoformat(FILTER_START_DATE_STR.replace("Z", "+00:00"))
except Exception as e:
    print(Fore.RED + "❌ Invalid FILTER_START_DATE in .env. Use format: 2025-07-26T00:00:00Z")
    logging.error("Invalid FILTER_START_DATE in .env")
    exit(1)

def read_last_timestamp():
    if os.path.exists("last_time.txt"):
        with open("last_time.txt", "r") as f:
            ts = f.read().strip()
            try:
                return datetime.fromisoformat(ts)
            except:
                return None
    return None

def save_last_timestamp(dt):
    with open("last_time.txt", "w") as f:
        f.write(dt.isoformat())

def fetch_email_after_date(sender_email, after_utc):
    print(Fore.CYAN + "🔄 Connecting to Gmail...")
    imap = imaplib.IMAP4_SSL("imap.gmail.com")
    imap.login(GMAIL_USER, GMAIL_PASS)
    imap.select("inbox")
    status, messages = imap.search(None, f'FROM "{sender_email}"')
    email_ids = messages[0].split()

    for eid in reversed(email_ids):  # newest first
        status, msg_data = imap.fetch(eid, "(RFC822)")
        if status != 'OK':
            continue
        msg = email.message_from_bytes(msg_data[0][1])
        try:
            email_date = parsedate_to_datetime(msg["Date"]).astimezone(timezone.utc)
        except Exception:
            continue
        if email_date <= after_utc:
            continue

        body = ""
        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_type() == "text/plain":
                    body += part.get_payload(decode=True).decode(errors="ignore")
        else:
            body = msg.get_payload(decode=True).decode(errors="ignore")

        imap.logout()
        return body, email_date

    imap.logout()
    return None, after_utc

def parse_with_llm(email_body):
    prompt = f"""
You are an assistant. Extract the user's name, consumed data %, and remaining data % from the email below. Output it in this format:

User | Consumed Data | Remaining Data
Mandar | 80% | 20%

If consumed data is more than 85%, also say:
"❗ Data consumption is high. Would you like to notify the user?"

Only return plain text table and optional warning.

Email:
---
{email_body}
---
"""
    print(Fore.MAGENTA + "🧠 Sending request to LLM...")
    response = requests.post(
        "http://localhost:11434/api/generate",
        json={
            "model": "gemma3",
            "prompt": prompt,
            "stream": False
        }
    )

    if response.status_code != 200:
        print(Fore.RED + f"❌ LLM request failed: {response.status_code}")
        logging.error(f"LLM request failed: {response.status_code} {response.text}")
        return ""

    return response.json().get("response", "")

def append_to_excel(llm_output, filename="output.xlsx"):
    rows = []
    for line in llm_output.strip().splitlines():
        if "|" in line and "User" not in line and "-----" not in line:
            cells = [cell.strip() for cell in line.split("|")]
            rows.append(cells)

    if not rows:
        print(Fore.YELLOW + "⚠️ No structured data found.")
        logging.warning("No structured data to append to Excel.")
        return

    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["User", "Consumed Data", "Remaining Data"])

    for row in rows:
        ws.append(row)

    wb.save(filename)
    print(Fore.GREEN + f"✅ Excel updated: {filename}")
    logging.info(f"Appended to Excel: {filename}")

def send_alert_email(to_email, subject, body):
    msg = EmailMessage()
    msg["From"] = GMAIL_USER
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.set_content(body)

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(GMAIL_USER, GMAIL_PASS)
            smtp.send_message(msg)
        print(Fore.GREEN + f"✅ Alert email sent to {to_email}")
        logging.info(f"Alert email sent to {to_email}")
    except Exception as e:
        print(Fore.RED + f"❌ Failed to send alert email: {e}")
        logging.error(f"Failed to send alert email: {e}")

def handle_warning_if_any(llm_output, sender_email):
    if "notify the user" in llm_output.lower():
        print(Fore.YELLOW + "❗ LLM suggests sending a notification to the user.")
        choice = input(Fore.CYAN + "📧 Do you want to send an email alert to the user? (yes/no): ").strip().lower()
        if choice == "yes":
            subject = "Data Usage Alert"
            body = "Hi,\n\nYour data consumption has exceeded 85%.\nPlease take necessary action.\n\nThanks"
            send_alert_email(sender_email, subject, body)
        else:
            print(Fore.CYAN + "🚫 No notification sent.")
            logging.info("User declined to send notification.")

# Main loop
if __name__ == "__main__":
    print(Fore.CYAN + "📧 Gmail Monitor + LLM + Excel + Alert (Running until Ctrl+C)")

    last_ts = read_last_timestamp()
    if last_ts is None:
        last_ts = FILTER_START_DATE

    print(Fore.BLUE + f"🕒 Monitoring inbox from {last_ts.isoformat()} UTC")
    logging.info(f"Started monitoring from {last_ts.isoformat()}")

    try:
        while True:
            email_body, email_time = fetch_email_after_date(TARGET_SENDER, last_ts)

            if email_body:
                print(Fore.GREEN + "📨 New Email Detected.")
                llm_output = parse_with_llm(email_body)
                print(Fore.WHITE + "🧠 LLM Output:\n", llm_output)

                append_to_excel(llm_output)
                handle_warning_if_any(llm_output, TARGET_SENDER)

                last_ts = email_time
                save_last_timestamp(last_ts)

            time.sleep(30)

    except KeyboardInterrupt:
        print(Fore.RED + "\n🛑 Program stopped by user.")
        logging.info("Program terminated by user.")
